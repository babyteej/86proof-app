"""
86 Proof Data Loader
─────────────────────────────────────────
Reads an uploaded 86 Proof Excel file and extracts the
relevant data tabs into clean Python dictionaries.

For v1, we read:
  - Menu Summary (cocktail pricing and pour cost)
  - Menu Engineering (sales mix, classification, performance)
  - Recipes (ingredient composition)
  - Ingredients (cost per unit, supplier)
  - Waste Log (waste events)

Variance is intentionally excluded — see scope decisions.
"""

import openpyxl


def safe_float(val):
    """Convert a value to float, returning None if it can't be converted."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def load_sheet(workbook, sheet_name):
    """Return list of row-dicts for a given sheet, or None if not found."""
    if sheet_name not in workbook.sheetnames:
        return None

    ws = workbook[sheet_name]
    all_rows = [r for r in ws.iter_rows(values_only=True)
                if any(v is not None for v in r)]

    if len(all_rows) < 2:
        return None

    # Find the actual header row (first row with 3+ string cells)
    header_idx = 0
    for i, row in enumerate(all_rows[:8]):
        string_count = sum(1 for v in row if isinstance(v, str) and len(str(v)) > 1)
        if string_count >= 3:
            header_idx = i
            break

    headers = [str(v).strip() if v is not None else f"col_{i}"
               for i, v in enumerate(all_rows[header_idx])]

    records = []
    for row in all_rows[header_idx + 1:]:
        rec = {headers[i]: row[i]
               for i in range(min(len(headers), len(row)))}
        if any(v is not None for v in rec.values()):
            records.append(rec)

    return records


def extract_menu_summary(records):
    """Extract cocktail-level pricing and pour cost data."""
    out = []
    for r in records:
        name = str(r.get("Cocktail", "")).strip()
        if not name or name.lower() in ("cocktail", "nan", ""):
            continue
        out.append({
            "name": name,
            "cogs": safe_float(r.get("COGS")),
            "menu_price": safe_float(r.get("Menu Price")),
            "pour_cost_pct": safe_float(r.get("Pour Cost %")),
            "margin": safe_float(r.get("Margin $")),
            "status": str(r.get("Status", "")).strip(),
        })
    return out


def extract_menu_engineering(records):
    """Extract sales performance and menu engineering classification."""
    out = []
    skip_keywords = ("cocktail", "nan", "", "total", "star", "plowhorse",
                     "puzzle", "dog", "average", "summary", "class")
    for r in records:
        name = str(r.get("Cocktail", "")).strip()
        if not name or name.lower().strip("⭐🐴🧩🐕 ") in skip_keywords:
            continue
        if r.get("Units Sold (period)") is None and r.get("Classification") is None:
            continue
        out.append({
            "name": name,
            "classification": str(r.get("Classification", "")).strip(),
            "units_sold": safe_float(r.get("Units Sold (period)")),
            "sales_mix_pct": safe_float(r.get("Sales Mix %")),
            "contrib_margin": safe_float(r.get("Contribution Margin $")),
            "menu_price": safe_float(r.get("Menu Price")),
            "cogs": safe_float(r.get("COGS")),
        })
    return out


def extract_recipes(records):
    """Extract cocktail recipes — which ingredients in which amounts."""
    out = []
    for r in records:
        name = str(r.get("Cocktail", "")).strip()
        if not name or name.lower() in ("cocktail", "nan", ""):
            continue

        # Recipes have up to 6 ingredient slots
        ingredients = []
        for i in range(1, 7):
            ing_name = r.get(f"Ing {i}")
            ing_amt = r.get(f"Amt {i} (ml)")
            if ing_name and str(ing_name).strip():
                ingredients.append({
                    "name": str(ing_name).strip(),
                    "amount_ml": safe_float(ing_amt),
                })

        out.append({
            "name": name,
            "ingredients": ingredients,
            "total_cogs": safe_float(r.get("Total COGS")),
            "menu_price": safe_float(r.get("Menu Price")),
            "pour_cost_pct": safe_float(r.get("Actual Pour Cost %")),
        })
    return out


def extract_ingredients(records):
    """Extract ingredient-level cost and supplier data."""
    out = []
    for r in records:
        name = str(r.get("Ingredient", "")).strip()
        if not name or name.lower() in ("ingredient", "nan", ""):
            continue
        out.append({
            "name": name,
            "category": str(r.get("Category", "")).strip(),
            "supplier": str(r.get("Supplier", "")).strip(),
            "cost_per_unit": safe_float(r.get("Cost per Nominal Unit")),
            "yield_pct": safe_float(r.get("Yield %")),
        })
    return out


def extract_waste(records):
    """Extract waste events."""
    out = []
    for r in records:
        item = str(r.get("Item Wasted", "")).strip()
        if not item or item.lower() in ("item wasted", "nan", ""):
            continue
        date_val = str(r.get("Date", ""))
        if not any(c.isdigit() for c in date_val):
            continue
        cost = safe_float(r.get("Cost $"))
        if cost and cost > 0:
            out.append({
                "date": str(r.get("Date", ""))[:10],
                "item": item,
                "cost": cost,
                "reason": str(r.get("Reason", "")).strip(),
                "notes": str(r.get("Notes", "")).strip(),
            })
    return out


def load_all_data(file_path):
    """
    Main entry point. Reads an 86 Proof Excel file and returns
    a dictionary containing all the extracted data tabs.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    data = {}

    tab_map = [
        ("menu_summary", "Menu Summary", extract_menu_summary),
        ("menu_engineering", "Menu Engineering", extract_menu_engineering),
        ("recipes", "Recipes", extract_recipes),
        ("ingredients", "Ingredients", extract_ingredients),
        ("waste", "Waste Log", extract_waste),
    ]

    for key, tab_name, extractor in tab_map:
        records = load_sheet(wb, tab_name)
        if records is not None:
            extracted = extractor(records)
            if extracted:
                data[key] = extracted

    return data